"""TO-DO: Write a description of what this XBlock is."""

import pkg_resources
import urllib
import os
import json

from collections import OrderedDict
from functools import partial

from xblock.core import XBlock
from xblock.fields import Scope, Integer, String, Float
from xblock.fragment import Fragment
from django.template import Context, Template

from openpyxl import load_workbook
from webob.response import Response
from xmodule.contentstore.content import StaticContent
from xmodule.contentstore.django import contentstore

from xblock_django.mixins import FileUploadMixin


class NeighborhoodDynamicsXBlock(XBlock, FileUploadMixin):
    """
    TO-DO: document what your XBlock does.
    """
    json_data = String(help="JSON data from excel file", default=None, scope=Scope.content)

    display_name = String(display_name="Display Name",
                          default="Neighborhood Dynamics",
                          scope=Scope.settings,
                          help="Name of Xblock.")

    san_felipe_lower = Float(display_name="San Felipe Lower Bound",
                             default=1.0,
                             scope=Scope.settings,
                             help="Lower bound")

    san_felipe_upper = Float(display_name="San Felipe Upper Bound",
                             default=1.0,
                             scope=Scope.settings,
                             help="Upper bound")

    santa_ana_lower = Float(display_name="Santa Ana Lower Bound",
                            default=1.0,
                            scope=Scope.settings,
                            help="Lower bound")

    santa_ana_upper = Float(display_name="Santa Ana Upper Bound",
                            default=1.0,
                            scope=Scope.settings,
                            help="Upper bound")

    el_chorillo_lower = Float(display_name="El Chorillo Lower Bound",
                              default=1.0,
                              scope=Scope.settings,
                              help="Lower bound")

    el_chorillo_upper = Float(display_name="El Chorillo Upper Bound",
                              default=1.0,
                              scope=Scope.settings,
                              help="Upper bound")

    san_felipe_1 = Float(display_name="San Felipe 1990 to 2000",
                         default=1.0,
                         scope=Scope.settings,
                         help="Percentage change")

    san_felipe_2 = Float(display_name="San Felipe 2000 to 2010",
                         default=1.0,
                         scope=Scope.settings,
                         help="Percentage change")

    san_felipe_3 = Float(display_name="San Felipe 1990 to 2010",
                         default=1.0,
                         scope=Scope.settings,
                         help="Percentage change")

    santa_ana_1 = Float(display_name="Santa Ana 1990 to 2000",
                        default=1.0,
                        scope=Scope.settings,
                        help="Percentage change")

    santa_ana_2 = Float(display_name="Santa Ana 2000 to 2010",
                        default=1.0,
                        scope=Scope.settings,
                        help="Percentage change")

    santa_ana_3 = Float(display_name="Santa Ana 1990 to 2010",
                        default=1.0,
                        scope=Scope.settings,
                        help="Percentage change")

    el_chorillo_1 = Float(display_name="El Chorillo 1990 to 2000",
                          default=1.0,
                          scope=Scope.settings,
                          help="Percentage change")

    el_chorillo_2 = Float(display_name="El Chorillo 2000 to 2010",
                          default=1.0,
                          scope=Scope.settings,
                          help="Percentage change")

    el_chorillo_3 = Float(display_name="El Chorillo 1990 to 2010",
                          default=1.0,
                          scope=Scope.settings,
                          help="Percentage change")

    task1_tries = Integer(
        default=0, scope=Scope.user_state,
        help="Counter of user tries", )

    task2_tries = Integer(
        default=0, scope=Scope.user_state,
        help="Counter of user tries", )

    def resource_string(self, path):
        """Handy helper for getting resources from our kit."""
        data = pkg_resources.resource_string(__name__, path)
        return data.decode("utf8")

    # TO-DO: change this view to display your data your own way.
    def student_view(self, context=None):
        """
        The primary view of the NeighborhoodDynamicsXBlock, shown to students
        when viewing courses.
        """

        template_str = self.resource_string("static/html/neighborhood_dynamics.html")
        template = Template(template_str)

        # parameters sent to browser for XBlock html page
        html = template.render(Context({
        }))

        frag = Fragment(html)
        frag.add_css_url(
            self.runtime.local_resource_url(
                self, 'public/css/neighborhood_dynamics.css'))
        frag.add_css(self.resource_string("static/css/multibar_charts.css"))
        frag.add_css(self.resource_string("static/css/nvd3.css"))
        frag.add_javascript(self.resource_string("static/js/src/d3.v3.js"))
        frag.add_javascript(self.resource_string("static/js/src/nvd3.js"))
        frag.add_javascript_url(self.runtime.local_resource_url(self, 'public/dist/bundle.js'))
        frag.add_javascript(self.resource_string("static/js/src/neighborhood_dynamics.js"))
        frag.initialize_js('NeighborhoodDynamicsXBlock', {
            'json_data': self.json_data
        })
        return frag

    def studio_view(self, context=None):
        """
        Create a fragment used to display the edit view in the Studio.
        """

        template_str = self.resource_string("static/html/studio_view.html")
        template = Template(template_str)
        html = template.render(Context({
            'display_name': self.display_name,
            'display_description': self.display_description,
            'san_felipe_lower': self.san_felipe_lower,
            'san_felipe_upper': self.san_felipe_upper,
            'santa_ana_lower': self.santa_ana_lower,
            'santa_ana_upper': self.santa_ana_upper,
            'el_chorillo_lower': self.el_chorillo_lower,
            'el_chorillo_upper': self.el_chorillo_upper,
            'san_felipe_1': self.san_felipe_1,
            'san_felipe_2': self.san_felipe_2,
            'san_felipe_3': self.san_felipe_3,
            'santa_ana_1': self.santa_ana_1,
            'santa_ana_2': self.santa_ana_2,
            'santa_ana_3': self.santa_ana_3,
            'el_chorillo_1': self.el_chorillo_1,
            'el_chorillo_2': self.el_chorillo_2,
            'el_chorillo_3': self.el_chorillo_3
        }))

        frag = Fragment(html.format(self=self))
        js_str = pkg_resources.resource_string(__name__, "static/js/src/studio_edit.js")
        frag.add_javascript(unicode(js_str))
        frag.initialize_js('StudioEdit')
        return frag

    @XBlock.handler
    def studio_submit(self, request, suffix=''):
        """
        Called when submitting the form in Studio.
        """
        data = request.POST
        self.display_name = data['display_name']
        self.display_description = data['display_description']
        self.san_felipe_lower = data['san_felipe_lower']
        self.san_felipe_upper = data['san_felipe_upper']
        self.santa_ana_lower = data['santa_ana_lower']
        self.santa_ana_upper = data['santa_ana_upper']
        self.el_chorillo_lower = data['el_chorillo_lower']
        self.el_chorillo_upper = data['el_chorillo_upper']
        self.san_felipe_1 = data['san_felipe_1']
        self.san_felipe_2 = data['san_felipe_2']
        self.san_felipe_3 = data['san_felipe_3']
        self.santa_ana_1 = data['santa_ana_1']
        self.santa_ana_2 = data['santa_ana_2']
        self.santa_ana_3 = data['santa_ana_3']
        self.el_chorillo_1 = data['el_chorillo_1']
        self.el_chorillo_2 = data['el_chorillo_2']
        self.el_chorillo_3 = data['el_chorillo_3']

        block_id = data['usage_id']
        if not isinstance(data['thumbnail'], basestring):
            upload = data['thumbnail']
            self.thumbnail_url = self.upload_to_s3('THUMBNAIL', upload.file, block_id, self.thumbnail_url)

        if not isinstance(data['excel'], basestring):
            upload = data['excel']

            # get workbook
            workbook = load_workbook(filename=upload.file, read_only=True)
            sheets = []
            # TODO: refactor!
            # this json will turn out looking bad. But refactoring will cause big changes on fronted
            # No time for this at the moment

            for worksheet in workbook:
                sheet = {
                    "name": worksheet.title,
                    "rows": []
                }
                if not ((worksheet.title == 'specs') or (worksheet.title == 'charts')):
                    for row in worksheet.iter_rows():
                        sheet_row = {
                            "key": None,
                            "values": []
                        }
                        cell_num = 0
                        # first row will be key for iteration, ex. "Country name"
                        for cell in row:
                            if cell_num is 0:
                                sheet_row["key"] = cell.value
                            else:
                                sheet_row["values"].append(cell.value)
                            cell_num += 1
                        sheet['rows'].append(sheet_row)
                    sheets.append(sheet)
                else:
                    # we will format differently sheets with specs
                    for row in worksheet.iter_rows():
                        _row = []
                        for cell in row:
                            _row.append(cell.value)
                        sheet['rows'].append(_row)
                    sheets.append(sheet)
            self.json_data = json.dumps(sheets)

        return Response(json_body={
            'result': "success"
        })

    @XBlock.json_handler
    def submit_task1(self, data, suffix=''):
        self.task1_tries = self.task1_tries + 1;
        san_felipe = data['san_felipe_lower'] and data['san_felipe_upper'] and (
        self.san_felipe_lower == float(data['san_felipe_lower'])) and (
                     self.san_felipe_upper == float(data['san_felipe_upper']))
        santa_ana = data['santa_ana_lower'] and data['santa_ana_upper'] and (
        self.santa_ana_lower == float(data['santa_ana_lower'])) and (
                    self.santa_ana_upper == float(data['santa_ana_upper']))
        el_chorillo = data['el_chorillo_lower'] and data['el_chorillo_upper'] and (
        self.el_chorillo_lower == float(data['el_chorillo_lower'])) and (
                      self.el_chorillo_upper == float(data['el_chorillo_upper']))

        return {
            'result': 'success',
            'data': {
                'san_felipe': san_felipe,
                'santa_ana': santa_ana,
                'el_chorillo': el_chorillo,
                'tries': self.task1_tries
            }
        }

    @XBlock.json_handler
    def submit_task2(self, data, suffix=''):
        self.task2_tries = self.task2_tries + 1;
        san_felipe1 = data['san_felipe_1'] and data['san_felipe_1'] and (
        self.san_felipe_1 == float(data['san_felipe_1'])) and (self.san_felipe_1 == float(data['san_felipe_1']))
        san_felipe2 = data['san_felipe_2'] and data['san_felipe_2'] and (
        self.san_felipe_2 == float(data['san_felipe_2'])) and (self.san_felipe_2 == float(data['san_felipe_2']))
        san_felipe3 = data['san_felipe_3'] and data['san_felipe_3'] and (
        self.san_felipe_3 == float(data['san_felipe_3'])) and (self.san_felipe_3 == float(data['san_felipe_3']))
        santa_ana1 = data['santa_ana_1'] and data['santa_ana_1'] and (
        self.santa_ana_1 == float(data['santa_ana_1'])) and (self.santa_ana_1 == float(data['santa_ana_1']))
        santa_ana2 = data['santa_ana_2'] and data['santa_ana_2'] and (
        self.santa_ana_2 == float(data['santa_ana_2'])) and (self.santa_ana_2 == float(data['santa_ana_2']))
        santa_ana3 = data['santa_ana_3'] and data['santa_ana_3'] and (
        self.santa_ana_3 == float(data['santa_ana_3'])) and (self.santa_ana_3 == float(data['santa_ana_3']))
        el_chorillo1 = data['el_chorillo_1'] and data['el_chorillo_1'] and (
        self.el_chorillo_1 == float(data['el_chorillo_1'])) and (self.el_chorillo_1 == float(data['el_chorillo_1']))
        el_chorillo2 = data['el_chorillo_2'] and data['el_chorillo_2'] and (
        self.el_chorillo_2 == float(data['el_chorillo_2'])) and (self.el_chorillo_2 == float(data['el_chorillo_2']))
        el_chorillo3 = data['el_chorillo_3'] and data['el_chorillo_3'] and (
        self.el_chorillo_3 == float(data['el_chorillo_3'])) and (self.el_chorillo_3 == float(data['el_chorillo_3']))

        return {
            'result': 'success',
            'data': {
                'san_felipe1': san_felipe1,
                'san_felipe2': san_felipe2,
                'san_felipe3': san_felipe3,
                'santa_ana1': santa_ana1,
                'santa_ana2': santa_ana2,
                'santa_ana3': santa_ana3,
                'el_chorillo1': el_chorillo1,
                'el_chorillo2': el_chorillo2,
                'el_chorillo3': el_chorillo3,
                'tries': self.task2_tries
            }
        }

    def _file_storage_name(self, filename):
        # pylint: disable=no-member
        """
        Get file path of storage.
        """
        path = (
            '{loc.block_type}/{loc.block_id}'
            '/{filename}'.format(
                loc=self.location,
                filename=filename
            )
        )
        return path

    # TO-DO: change this to create the scenarios you'd like to see in the
    # workbench while developing your XBlock.
    @staticmethod
    def workbench_scenarios():
        """A canned scenario for display in the workbench."""
        return [
            ("NeighborhoodDynamicsXBlock",
             """<neighborhood_dynamics/>
             """),
            ("Multiple NeighborhoodDynamicsXBlock",
             """<vertical_demo>
                <neighborhood_dynamics/>
                <neighborhood_dynamics/>
                <neighborhood_dynamics/>
                </vertical_demo>
             """),
        ]
